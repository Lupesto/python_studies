from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws1 = wb.create_sheet("Panilha1") # insert at the end (default)


ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA" #cor na planilha
ws['A4'] = 4
ws1['A4'] = 4

wb.save('balances.xlsx')
